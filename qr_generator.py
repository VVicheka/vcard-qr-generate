"""Reusable QR + vCard generation logic.

Used by both the CLI script and the Flask web app.
"""

import io
import re
import zipfile
from openpyxl import load_workbook
from PIL import Image
import qrcode
from qrcode.image.styledpil import StyledPilImage
from qrcode.image.styles.moduledrawers.pil import StyledPilQRModuleDrawer
from PIL import ImageDraw

# Supersampling factor for smooth circle edges (matches qrcode's own drawer).
_ANTIALIAS = 4


class SmallCircleModuleDrawer(StyledPilQRModuleDrawer):
    """Draws each module as a small centered circle.

    ``dot_ratio`` is the circle diameter relative to the module size: 1.0
    fills the whole module (touching neighbours), lower values give thinner
    dots with more white space around them. Thinner dots read more cleanly
    when printed (e.g. UV).
    """

    def __init__(self, dot_ratio=0.6):
        self.dot_ratio = dot_ratio

    def initialize(self, *args, **kwargs):
        super().initialize(*args, **kwargs)
        box_size = self.img.box_size
        fake_size = box_size * _ANTIALIAS
        diameter = max(1, int(fake_size * self.dot_ratio))
        offset = (fake_size - diameter) // 2
        self.circle = Image.new(
            self.img.mode, (fake_size, fake_size), self.img.color_mask.back_color
        )
        ImageDraw.Draw(self.circle).ellipse(
            (offset, offset, offset + diameter, offset + diameter),
            fill=self.img.paint_color,
        )
        self.circle = self.circle.resize((box_size, box_size), Image.LANCZOS)

    def drawrect(self, box, is_active: bool):
        if is_active:
            self.img._img.paste(self.circle, (box[0][0], box[0][1]))


# vCard field type -> how it renders in the vCard text
VCARD_RENDERERS = {
    "phone":   lambda label, val: f"TEL;TYPE=CELL:{val}",
    "email":   lambda label, val: f"EMAIL;TYPE=INTERNET:{val}",
    "url":     lambda label, val: f"URL;TYPE={label or 'Website'}:{val}",
    "note":    lambda label, val: f"NOTE:{val}",
    "title":   lambda label, val: f"TITLE:{val}",
    "org":     lambda label, val: f"ORG:{val}",
}


def build_vcard(first, last, fields):
    """Build a vCard string.

    fields: list of {"label": str, "value": str, "type": "phone|email|url|note"}
    """
    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"N:{last};{first};;;",
        f"FN:{first} {last}".strip(),
    ]
    for f in fields:
        val = (f.get("value") or "").strip()
        if not val:
            continue
        ftype = f.get("type", "url")
        label = (f.get("label") or "").strip()
        renderer = VCARD_RENDERERS.get(ftype, VCARD_RENDERERS["url"])
        lines.append(renderer(label, val))
    lines.append("END:VCARD")
    return "\n".join(lines)


def make_qr_with_logo(data, logo_bytes=None, qr_size=2000, logo_ratio=0.22):
    """Return a PIL Image of the QR with an optional centered logo.

    Error correction is kept as low as is safe so the QR has fewer modules
    (bigger squares) and stays scannable when printed small. A logo covers
    part of the code, so we bump the level up when one is present.
    """
    # Logo covers ~logo_ratio^2 of the area (~5% at 0.22). Q (25%) leaves
    # plenty of margin; M (15%) is fine with no logo and gives the cleanest QR.
    error_correction = (
        qrcode.constants.ERROR_CORRECT_Q if logo_bytes
        else qrcode.constants.ERROR_CORRECT_M
    )
    qr = qrcode.QRCode(
        version=None,
        error_correction=error_correction,
        box_size=20,
        border=4,  # QR spec minimum quiet zone; below 4 print scans often fail
    )
    qr.add_data(data)
    qr.make(fit=True)
    # Dot-style modules: each "on" module is drawn as a filled circle, which
    # holds up better when printed (e.g. UV) than tightly-packed squares.
    qr_img = qr.make_image(
        image_factory=StyledPilImage,
        module_drawer=SmallCircleModuleDrawer(dot_ratio=0.6),
        fill_color="black",
        back_color="white",
    ).convert("RGBA")

    natural = qr_img.size[0]
    scale = max(1, round(qr_size / natural))
    final_size = natural * scale
    if scale > 1:
        # LANCZOS keeps the dot edges smooth/round when scaling up,
        # whereas NEAREST would re-introduce blocky pixelation.
        qr_img = qr_img.resize((final_size, final_size), Image.LANCZOS)

    if logo_bytes:
        logo = Image.open(io.BytesIO(logo_bytes)).convert("RGBA")
        max_dim = int(final_size * logo_ratio)
        logo.thumbnail((max_dim, max_dim), Image.LANCZOS)
        logo_w, logo_h = logo.size

        pad = max(12, max(logo_w, logo_h) // 12)
        bg_w = logo_w + pad * 2
        bg_h = logo_h + pad * 2
        bg = Image.new("RGBA", (bg_w, bg_h), (255, 255, 255, 255))
        bg.paste(logo, (pad, pad), logo)

        pos = ((final_size - bg_w) // 2, (final_size - bg_h) // 2)
        qr_img.paste(bg, pos, bg)

    return qr_img.convert("RGB")


def qr_png_bytes(data, logo_bytes=None, qr_size=2000):
    img = make_qr_with_logo(data, logo_bytes=logo_bytes, qr_size=qr_size)
    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    buf.seek(0)
    return buf


def read_excel_headers(file_stream):
    wb = load_workbook(file_stream, data_only=True, read_only=True)
    ws = wb.active
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value).strip() if cell.value else "")
    return [h for h in headers if h]


def read_excel_rows(file_stream):
    """Return (headers, rows) where rows is list of dicts {header: value}."""
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        row = {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(raw) if i < len(headers)}
        rows.append(row)
    return headers, rows


_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def _safe_filename(name):
    name = _INVALID_FILENAME_CHARS.sub("", name).strip().strip(".")
    return name or "qr"


def build_zip_with_qr(rows, mapping, logo_bytes=None):
    """Build a ZIP archive of per-row PNG QR codes named by first/last.

    rows: list of dicts (from read_excel_rows)
    mapping: same shape as build_excel_with_qr expected.

    Returns BytesIO of the zip file.
    """
    first_col = mapping["first_col"]
    last_col = mapping["last_col"]

    out = io.BytesIO()
    seen = {}
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            first = row.get(first_col, "")
            last = row.get(last_col, "")

            vcard_fields = []
            for f in mapping["fields"]:
                val = row.get(f["col"], "")
                vcard_fields.append({"label": f.get("label") or f["col"], "value": val, "type": f["type"]})

            vcard = build_vcard(first, last, vcard_fields)
            png = qr_png_bytes(vcard, logo_bytes=logo_bytes)

            base = _safe_filename(f"{first}_{last}".strip("_"))
            count = seen.get(base, 0) + 1
            seen[base] = count
            filename = f"{base}.png" if count == 1 else f"{base}_{count}.png"
            zf.writestr(filename, png.read())

    out.seek(0)
    return out
